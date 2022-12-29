import zipfile
import os.path
import glob
from os.path import basename
from PyPDF2 import PdfReader
from openpyxl import load_workbook
import csv

path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "resources")  # файлы для архива
path_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), "files")
path_zip = os.path.join(path_file, "myzip.zip")  # сам архив
file_dir = os.listdir(path)

# with zipfile.ZipFile(path_zip, mode="w", compression=zipfile.ZIP_DEFLATED) as archive:
#     for file in file_dir:
#         add_file = os.path.join(path, file)
#         archive.write(add_file, basename(add_file))
#
# with zipfile.ZipFile(path_zip) as archive:
#     pdf_file = archive.extract("test_PDF.pdf")
#     reader = PdfReader(pdf_file)
#     page = reader.pages[0]
#     text = page.extract_text()
# assert 'Тестовый PDF' in text, f"Expected result: {'Тестовый PDF'}; actual result: {text}"
# os.remove(pdf_file)


# with zipfile.ZipFile(path_zip) as archive:
#     xf = archive.extract("excel_XLSX.xlsx")
#     xlsx_file = load_workbook(xf)
#     sheet = xlsx_file.active
#     # print(sheet.cell(row=2, column=2).value)
#     assert sheet.cell(row=2, column=2).value == "Dulce", f"Expected result: {'Dulce'}, "f"actual result: {sheet.cell(row=2, column=2).value}"
#     os.remove(xf)

with zipfile.ZipFile(path_zip) as archive:
    csv_file = archive.extract("test_CSV.csv")
with open(csv_file) as csvfile:
    csvfile = csv.reader(csvfile)
# list_csv = []
for r in csvfile:
    print(r)
#
# with open('resources/test_CSV.csv') as csvfile:
#     csvfile = csv.reader(csvfile)
#     for r in csvfile:
#         print(r)


#     text = "".join(r).replace(";", " ")
# list_csv.append(text)
# assert list_csv[4] == "Test python 123456 Empire State", f"Expected result: {'Test python 123456 Empire State'}, "f"actual result: {list_csv[4]}"
# os.remove(csv_file)



# with zipfile.ZipFile(path_zip, 'a') as zip:
#     for file in zip.namelist():
#         print(file)


# @pytest.fixture()
# def clear_dir():
#     all_files = os.path.join(file_dir, '*.*')
#     for file in glob.glob(all_files):
#         os.remove(file)



